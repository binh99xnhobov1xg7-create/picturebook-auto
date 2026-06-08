"""一次性脚本：把三个官方 S&S Excel 快照成 references/syllabus/syllabus.json。

运行期由 syllabus.py 读取该 JSON（可入库、不依赖 ~/下载 目录）。

兼容两套列模式：
  • L3-6（Level 3-6 S&S.xlsx）：Book No / 虚构非虚构 / 课文标题 / 句型 / 核心词 /
    定表词1-5 / 正文 / Lexile / CEFR / Phonics Rule|Word Formation / Reading Strategy /
    Reading Skill / Questions(Type) / Reading Report Questions / Graphic Organizer + 描述 + 使用建议
  • L0-2（S&S绘本大纲内容Level 0-Level 2.xlsx 的 Level 0/1/2 表）：Title / Genre / Reader /
    Vocabulary / Mastery / Exposure / Sentence Frames / Sor / Reading Strategy / Reading Skill /
    Questions / Answers / Oral Prompts / Extension / Phonics / Syntax / Morphology /
    Comprehension Tier / Cognitive Demand

用法：py scripts/_build_syllabus_snapshot.py
"""
from __future__ import annotations

import glob
import json
import re
from pathlib import Path

REPO = Path(__file__).resolve().parents[1]
OUT = REPO / "references" / "syllabus" / "syllabus.json"
DL = Path.home() / "下载"

# 官方 S&S 大纲固定存放点（用户 2026-06-08 规整后的权威搜索路径）：
#   ~/下载/VIPKID/大纲/  ← 以后所有 S&S Excel 都放这里，优先从此处取最新版。
# 兜底仍递归扫描整个 ~/下载（兼容历史/临时放置）。
SEARCH_DIRS = [
    DL / "VIPKID" / "大纲",
    DL / "VIPKID",
    DL,
]


def _scan_xlsx() -> dict[str, str]:
    """递归收集候选 S&S Excel：{文件名: 绝对路径}。

    同名文件出现在多处时，优先级按 SEARCH_DIRS 顺序（VIPKID/大纲 最高），
    其次取修改时间最新的，确保命中用户最新整理的权威大纲。
    """
    found: dict[str, tuple[int, float, str]] = {}  # name -> (优先级, mtime, path)
    for rank, base in enumerate(SEARCH_DIRS):
        if not base.exists():
            continue
        for p in base.glob("*.xlsx"):
            if p.name.startswith("~$"):  # 跳过 Excel 临时锁文件
                continue
            mtime = p.stat().st_mtime
            prev = found.get(p.name)
            if prev is None or rank < prev[0] or (rank == prev[0] and mtime > prev[1]):
                found[p.name] = (rank, mtime, str(p))
    return {name: meta[2] for name, meta in found.items()}


def _norm_title(s: str) -> str:
    """标题归一化键：小写、去标点空白，便于跨来源模糊匹配。"""
    s = (s or "").strip().lower()
    s = s.replace("&", " and ")
    s = re.sub(r"[^a-z0-9]+", "", s)
    return s


def _clean(v) -> str:
    if v is None:
        return ""
    s = str(v).replace("\r", "")
    s = s.replace("\u200b", "").replace("\ufeff", "")  # 零宽空格/BOM
    return s.strip()


def _clean_lexile(s: str) -> str:
    s = (s or "").replace("\n", " ")
    s = re.sub(r"(?i)lexile\s*range", "", s)
    return re.sub(r"\s+", " ", s).strip()


def _is_real_word(w: str) -> bool:
    """过滤词表里的数字/单字符噪音。"""
    w = (w or "").strip()
    return len(w) >= 2 and not w.replace(".", "").isdigit()


def _hmap(header_row) -> dict[str, int]:
    """{规范化表头: 列号}。规范化=小写去空白。"""
    m: dict[str, int] = {}
    for i, c in enumerate(header_row):
        key = _clean(c).lower()
        if key:
            m[key] = i
    return m


def _find(hmap: dict[str, int], *needles, exact: str | None = None) -> int:
    """按子串找列号；exact 优先精确匹配（去空白小写）。"""
    if exact is not None:
        ex = exact.lower().replace(" ", "")
        for k, i in hmap.items():
            if k.replace(" ", "") == ex:
                return i
    for k, i in hmap.items():
        if all(n.lower() in k for n in needles):
            return i
    return -1


def _get(row, idx) -> str:
    if idx < 0 or idx >= len(row):
        return ""
    return _clean(row[idx])


def _genre(raw: str) -> str:
    s = (raw or "").lower()
    if "非虚构" in raw or "nonfiction" in s or "non-fiction" in s or "informational" in s:
        return "nonfiction"
    if "虚构" in raw or "fiction" in s:
        return "fiction"
    return ""


# ============================================================
#  L3-6 表解析
# ============================================================
def parse_l36(ws, level: str) -> dict[str, dict]:
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    if not rows:
        return {}
    h = _hmap(rows[0])
    c_bookno = _find(h, "book no")
    c_genre = _find(h, "虚构")
    c_title = _find(h, "课文标题")
    c_pattern = _find(h, "句型")
    c_example = _find(h, "文中原句")
    c_focus = _find(h, "教学重点")
    c_text7 = _find(h, "课文正文调整")
    c_pure = _find(h, "纯正文")
    c_words = _find(h, "总字数")
    c_lex = _find(h, exact="Lexile")
    c_cefr = _find(h, exact="CEFR")
    c_phonics = _find(h, "phonics rule")
    if c_phonics < 0:
        c_phonics = _find(h, "word formation")
    c_decode_label = "Word Formation" if _find(h, "word formation") >= 0 else "Phonics Rule"
    c_phex = _find(h, "example words")
    c_strategy = _find(h, "reading strategy")
    c_skill = _find(h, "reading skill")
    c_qtype = _find(h, "questions")
    c_rr = _find(h, "reading report")
    c_go = _find(h, exact="Graphic Organizer")
    c_go_desc = _find(h, "graphic organizer", "描述")
    c_go_use = _find(h, "使用建议")
    c_core = _find(h, "核心词汇")   # 单列「核心词汇」（Level 3 新版仅此列，无定表词 1-5）
    # 定表词 1-5
    vocab_cols = []
    for n in range(1, 6):
        cw = _find(h, exact=f"定表词{n}")
        cl = _find(h, exact=f"单词级别{n}")
        ce = _find(h, exact=f"例句{n}")
        if cw >= 0:
            vocab_cols.append((cw, cl, ce))

    out: dict[str, dict] = {}
    for row in rows[1:]:
        title = _get(row, c_title)
        if not title:
            continue
        vocab = []
        for cw, cl, ce in vocab_cols:
            w = _get(row, cw)
            if not _is_real_word(w):
                continue
            vocab.append({"word": w, "level": _get(row, cl), "example": _get(row, ce)})
        # 回退：无「定表词」明细时（Level 3 新版），拆「核心词汇」单列成词表
        core_raw = _get(row, c_core)
        if not vocab and core_raw:
            for w in re.split(r"[,，、;；/]+", core_raw):
                w = w.strip()
                if _is_real_word(w):
                    vocab.append({"word": w, "level": "", "example": ""})
        booknum = _get(row, c_bookno) if c_bookno >= 0 else ""
        # Excel 数字常被读成 "1.0" → 规整成整型字符串
        if booknum:
            try:
                booknum = str(int(float(booknum)))
            except Exception:
                booknum = str(booknum).strip()
        entry = {
            "level": level,
            "title": title,
            "book_number": booknum,
            "genre": _genre(_get(row, c_genre)),
            "sentence_pattern": _get(row, c_pattern),
            "example_sentence": _get(row, c_example),
            "teaching_focus": _get(row, c_focus),
            "core_vocab": vocab,
            "core_vocab_raw": core_raw,
            "text_7page": _get(row, c_text7),
            "pure_text": _get(row, c_pure),
            "word_count": _get(row, c_words),
            "lexile": _clean_lexile(_get(row, c_lex)),
            "cefr": _get(row, c_cefr),
            "decoding_label": c_decode_label,
            "phonics_rule": _get(row, c_phonics),
            "phonics_examples": _get(row, c_phex),
            "reading_strategy": _get(row, c_strategy),
            "reading_skill": _get(row, c_skill),
            "questions_type": _get(row, c_qtype),
            "rr_questions": _get(row, c_rr),
            "graphic_organizer": _get(row, c_go),
            "go_description": _get(row, c_go_desc),
            "go_usage": _get(row, c_go_use),
        }
        # 重名书（如 L3 「Homes Around the World」#18 与 #56）按书号去重，不再互相覆盖：
        #   首次用纯标题键（保证 syllabus.match 按标题仍命中第一本）；重复则挂 标题#书号 键。
        base_key = f"{level}::{_norm_title(title)}"
        key = base_key if base_key not in out else f"{base_key}#{booknum or len(out)}"
        out[key] = {k: v for k, v in entry.items() if v not in ("", [], None)}
        out[key]["level"] = level
        out[key]["title"] = title
        if booknum:
            out[key]["book_number"] = booknum
    return out


# ============================================================
#  L0-2 表解析
# ============================================================
def parse_l02(ws, level: str) -> dict[str, dict]:
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    if not rows:
        return {}
    h = _hmap(rows[0])
    c_title = _find(h, exact="Title")
    c_genre = _find(h, exact="Genre")
    c_reader = _find(h, exact="Reader")
    c_words = _find(h, "word count")
    c_obj = _find(h, exact="Objective")
    c_vocab = _find(h, "vocabulary")
    c_mastery = _find(h, exact="Mastery")
    c_exposure = _find(h, exact="Exposure")
    c_frames = _find(h, "sentence frames")
    c_sor = _find(h, exact="Sor")
    c_strategy = _find(h, "reading strategy")
    c_skill = _find(h, "reading skill")
    c_q = _find(h, "questions")
    c_ans = _find(h, exact="Answers")
    c_oral = _find(h, "oral prompts")
    c_ext = _find(h, "extension")
    c_lex = _find(h, "lexile")
    c_cefr = _find(h, "cefr")
    c_phonics = _find(h, "phonics focus")
    c_syntax = _find(h, "syntax focus")
    c_morph = _find(h, "morphology")
    c_tier = _find(h, "comprehension tier")
    c_demand = _find(h, "cognitive demand")

    def _splitw(s: str) -> list[str]:
        return [w.strip() for w in re.split(r"[,，、;；/]+", s or "") if w.strip()]

    out: dict[str, dict] = {}
    for row in rows[1:]:
        title = _get(row, c_title)
        if not title:
            continue
        entry = {
            "level": level,
            "title": title,
            "genre": _genre(_get(row, c_genre)) or _get(row, c_genre),
            "reader_text": _get(row, c_reader),
            "word_count": _get(row, c_words),
            "objective": _get(row, c_obj),
            "core_vocab_raw": _get(row, c_vocab),
            "vocab_mastery": _splitw(_get(row, c_mastery)),
            "vocab_exposure": _splitw(_get(row, c_exposure)),
            "sentence_frames": _get(row, c_frames),
            "sor": _get(row, c_sor),
            "reading_strategy": _get(row, c_strategy),
            "reading_skill": _get(row, c_skill),
            "questions_type": _get(row, c_q),
            "answers": _get(row, c_ans),
            "oral_prompts": _get(row, c_oral),
            "extension": _get(row, c_ext),
            "lexile": _clean_lexile(_get(row, c_lex)),
            "cefr": _get(row, c_cefr),
            "phonics_rule": _get(row, c_phonics),
            "syntax_focus": _get(row, c_syntax),
            "morphology_focus": _get(row, c_morph),
            "comprehension_tier": _get(row, c_tier),
            "cognitive_demand": _get(row, c_demand),
        }
        key = f"{level}::{_norm_title(title)}"
        out[key] = {k: v for k, v in entry.items() if v not in ("", [], None)}
        out[key]["level"] = level
        out[key]["title"] = title
    return out


# ============================================================
#  SoR 分级策略表（Level 0/1/2 的 before/during/after）
# ============================================================
def parse_sor_sheet(ws) -> dict[str, dict]:
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    sor: dict[str, dict] = {}
    cur: str | None = None
    cols = {"before": 1, "during": 2, "after": 4}

    def first_line(cell: str) -> str:
        for ln in (cell or "").replace("\r", "").split("\n"):
            ln = ln.strip()
            if ln and "SoR Component" not in ln:
                return re.sub(r"^\d+\.\s*", "", ln).strip()
        return ""

    for row in rows:
        joined = " ".join(_clean(c) for c in row if c is not None)
        m = re.search(r"LEVEL\s*(\d)\s*READERS", joined.upper())
        if m:
            cur = m.group(1)
            sor.setdefault(cur, {"before": [], "during": [], "after": []})
            continue
        # 表头行（含 BEFORE READING 等）也确定 after 列位置
        if cur is None:
            continue
        for phase, ci in cols.items():
            txt = first_line(_get(row, ci))
            if txt and txt.upper() not in ("BEFORE READING (SOR-ALIGNED)", "DURING READING (SOR-ALIGNED)",
                                           "AFTER READING (SOR-ALIGNED)") and "READING (SOR" not in txt.upper():
                if txt not in sor[cur][phase]:
                    sor[cur][phase].append(txt)
    # 去掉空级别
    return {k: v for k, v in sor.items() if any(v.values())}


def main() -> None:
    import openpyxl

    files = _scan_xlsx()
    f_l36 = next((p for n, p in files.items() if "3-6" in n), None)
    f_l02 = next((p for n, p in files.items() if "0-Level 2" in n or "0-Level2" in n or "Level 0-Level" in n), None)

    books: dict[str, dict] = {}
    sor_strategies: dict[str, dict] = {}

    if f_l36:
        wb = openpyxl.load_workbook(f_l36, data_only=True)
        for lvl in ("3", "4", "5", "6"):
            name = f"Level {lvl}"
            if name in wb.sheetnames:
                books.update(parse_l36(wb[name], lvl))
        print("L3-6 books:", len([k for k in books]))

    if f_l02:
        wb = openpyxl.load_workbook(f_l02, data_only=True)
        for lvl in ("0", "1", "2"):
            name = f"Level {lvl}"
            if name in wb.sheetnames:
                books.update(parse_l02(wb[name], lvl))
        if "SoR" in wb.sheetnames:
            sor_strategies = parse_sor_sheet(wb["SoR"])
        print("with L0-2 books total:", len(books), "| SoR levels:", list(sor_strategies))

    OUT.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "_meta": {
            "source_l36": Path(f_l36).name if f_l36 else None,
            "source_l02": Path(f_l02).name if f_l02 else None,
            "book_count": len(books),
        },
        "books": books,
        "sor_strategies": sor_strategies,
    }
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=1), encoding="utf-8")
    print("WROTE", OUT, "(", OUT.stat().st_size, "bytes )")


if __name__ == "__main__":
    main()
