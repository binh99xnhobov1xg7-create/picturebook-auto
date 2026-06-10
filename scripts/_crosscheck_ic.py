# -*- coding: utf-8 -*-
"""Cross-check Timeline Book Club IC vs S&S for Level 3-4."""
from __future__ import annotations

import re
import shutil
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parent.parent
TIMELINE = Path(r"c:\Users\Jered\下载\Timeline -Dino Reading精读绘本课程整体排期.xlsx")
SS = Path(r"c:\Users\Jered\下载\VIPKID\大纲\Level 3-6  S&S.xlsx")
OUT_MD = ROOT / "docs" / "L3-L4_IC_交叉核对报告.md"
OUT_XLSX = ROOT / "outputs" / "_framework" / "L3-L4_IC待补标记.xlsx"
OUT_TIMELINE = ROOT / "outputs" / "_framework" / "Timeline_L3-L4_IC待制作标记.xlsx"

IP_NAMES = [
    "Mia", "Tommy", "Anna", "Teacher Kim", "Kim", "Winnie", "Dino",
    "Grandma", "Grandpa", "Mom", "Dad", "Uncle", "Cousin", "Jack", "Bob", "Nina",
]

EMPTY_IC = {"", "—", "-", "–", "N/A", "n/a", "NA", "无", "待定", "TBD", "tbd", "None"}


def ic_is_empty(ic_val) -> bool:
    if ic_val is None:
        return True
    s = str(ic_val).strip()
    return not s or s in EMPTY_IC


def story_ip_names(text: str) -> list[str]:
    if not text:
        return []
    found = []
    for name in IP_NAMES:
        if re.search(rf"\b{re.escape(name)}\b", text, re.I):
            found.append(name)
    seen: set[str] = set()
    out = []
    for n in found:
        k = n.lower()
        if k not in seen:
            seen.add(k)
            out.append(n)
    return out


def titles_match(a: str, b: str) -> bool:
    if not a or not b:
        return True
    al, bl = a.lower().strip(), b.lower().strip()
    return al == bl or al in bl or bl in al


def load_ss(level: str) -> dict[int, dict]:
    wb = openpyxl.load_workbook(SS, data_only=True)
    ws = wb[level]
    title_col = 4 if level == "Level 3" else 3
    theme_col = 3 if level == "Level 3" else None
    genre_col = 2

    story_col: int | None = 11 if level == "Level 3" else None
    if level == "Level 4":
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        for i, h in enumerate(headers):
            hs = str(h or "")
            if "不含 Page" in hs or "纯正文" in hs or "课文正文" in hs:
                story_col = i + 1
                break
        if not story_col:
            best_c, best_len = 3, 0
            for c in range(1, ws.max_column + 1):
                v = ws.cell(2, c).value
                if v and len(str(v)) > best_len:
                    best_len = len(str(v))
                    best_c = c
            story_col = best_c

    books: dict[int, dict] = {}
    for r in range(2, ws.max_row + 1):
        raw_no = ws.cell(r, 1).value
        if raw_no is None:
            continue
        try:
            no = int(float(raw_no))
        except (TypeError, ValueError):
            continue
        title = ws.cell(r, title_col).value
        story = ws.cell(r, story_col).value if story_col else ""
        theme = ws.cell(r, theme_col).value if theme_col else ""
        genre = ws.cell(r, genre_col).value
        books[no] = {
            "title": str(title or "").strip(),
            "theme": str(theme or "").strip(),
            "genre": str(genre or "").strip(),
            "story": str(story or ""),
            "ips_in_story": story_ip_names(str(story or "")),
        }
    wb.close()
    return books


def load_timeline(level: str) -> list[dict]:
    wb = openpyxl.load_workbook(TIMELINE, data_only=True)
    ws = wb[level]
    rows = []
    for r in range(4, ws.max_row + 1):
        raw_no = ws.cell(r, 1).value
        if raw_no is None:
            continue
        try:
            no = int(float(raw_no))
        except (TypeError, ValueError):
            continue
        title_tl = ws.cell(r, 2).value
        ic = ws.cell(r, 3).value
        if not str(title_tl or "").strip() and ic_is_empty(ic):
            continue
        ppt = ws.cell(r, 4).value
        rows.append({
            "book_no": no,
            "title_timeline": str(title_tl or "").strip(),
            "ic": ic,
            "ic_empty": ic_is_empty(ic),
            "ppt_status": str(ppt or "").strip(),
            "row": r,
        })
    wb.close()
    return rows


def mark_timeline_copy(pending: list[dict]) -> None:
    OUT_TIMELINE.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(TIMELINE, OUT_TIMELINE)
    wb = openpyxl.load_workbook(OUT_TIMELINE)
    fill = PatternFill("solid", fgColor="FEF3C7")
    bold = Font(bold=True, color="92400E")
    pending_by_level: dict[str, dict[int, dict]] = {"Level 3": {}, "Level 4": {}}
    for item in pending:
        lv = "Level 3" if item["level"] == "L3" else "Level 4"
        pending_by_level[lv][item["book_no"]] = item

    for level in ("Level 3", "Level 4"):
        ws = wb[level]
        ws.cell(3, 8, "制作状态")
        ws.cell(3, 8).font = bold
        for r in range(4, ws.max_row + 1):
            raw_no = ws.cell(r, 1).value
            try:
                no = int(float(raw_no))
            except (TypeError, ValueError):
                continue
            item = pending_by_level[level].get(no)
            if item:
                ws.cell(r, 3).fill = fill
                ws.cell(r, 8, "待分配IC / 未制作")
                ws.cell(r, 8).fill = fill
                ws.cell(r, 8).font = bold
    wb.save(OUT_TIMELINE)


def main() -> None:
    issues: list[dict] = []
    summary: list[tuple] = []

    for lv_key, lv_label in [("Level 3", "L3"), ("Level 4", "L4")]:
        ss = load_ss(lv_key)
        tl = load_timeline(lv_key)
        missing_ic: list[dict] = []
        title_mismatch: list[tuple] = []
        for item in tl:
            no = item["book_no"]
            ss_b = ss.get(no, {})
            ss_title = ss_b.get("title", "")
            tl_title = item["title_timeline"]
            title_ok = titles_match(ss_title, tl_title)

            if item["ic_empty"]:
                missing_ic.append({
                    "level": lv_label,
                    "book_no": no,
                    "title_ss": ss_title,
                    "title_timeline": tl_title,
                    "ic": "(空)",
                    "ips_in_story": ", ".join(ss_b.get("ips_in_story", [])) or "—",
                    "genre": ss_b.get("genre", ""),
                    "theme": ss_b.get("theme", ""),
                    "ppt_status": item.get("ppt_status", "") or "—",
                    "title_match": title_ok,
                })

            if ss_title and tl_title and not title_ok:
                title_mismatch.append((no, ss_title, tl_title))

        summary.append((lv_label, len(tl), len(missing_ic), len(title_mismatch)))
        issues.extend(missing_ic)

    examples = [
        ("L3", 30),
        ("L3", 42),
        ("L4", 6),
        ("L4", 9),
    ]
    example_checks = []
    for lv, no in examples:
        hit = next((x for x in issues if x["level"] == lv and x["book_no"] == no), None)
        ss = load_ss(f"Level {lv[1]}")
        tl_items = {t["book_no"]: t for t in load_timeline(f"Level {lv[1]}")}
        example_checks.append({
            "level": lv,
            "book_no": no,
            "in_missing_list": hit is not None,
            "ic": str(tl_items.get(no, {}).get("ic") or "(空)"),
            "title_ss": ss.get(no, {}).get("title", ""),
            "title_tl": tl_items.get(no, {}).get("title_timeline", ""),
        })

    OUT_MD.parent.mkdir(parents=True, exist_ok=True)
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)

    lines = [
        "# Level 3–4 IC 交叉核对报告",
        "",
        "数据源：",
        f"- 排期表（Book Club）：`{TIMELINE}` → Sheet **Level 3 / Level 4**，**C 列 IC = 制作负责人**",
        f"- 大纲 S&S：`{SS}` → 课文标题、体裁、正文 IP（供后续填 IC 参考）",
        "",
        "> **判定规则**：Timeline **IC 列为空** → 尚未分配负责人 / 尚未进入制作（PPT 列通常也为空）。",
        "> IC 列填的是人名（如张君蓝、冯尹），不是绘本角色名。",
        "",
        "## 汇总",
        "",
        "| 级别 | 排期表书目数 | IC 为空（待制作） | 已有 IC | 标题与 S&S 不一致 |",
        "|------|-------------|------------------|---------|-------------------|",
    ]
    for lv, total, miss, tm in summary:
        assigned = total - miss
        lines.append(f"| {lv} | {total} | **{miss}** | {assigned} | {tm} |")

    lines += [
        "",
        f"**合计待制作：{len(issues)} 本**（L3: {summary[0][2]} + L4: {summary[1][2]}）",
        "",
        "## 你举的例子（复核）",
        "",
        "| 级别 | Book# | S&S 标题 | Timeline IC | 是否在待制作清单 |",
        "|------|-------|---------|-------------|-----------------|",
    ]
    for ex in example_checks:
        lines.append(
            f"| {ex['level']} | {ex['book_no']} | {ex['title_ss'] or ex['title_tl']} | "
            f"{ex['ic']} | {'✅ 是' if ex['in_missing_list'] else '❌ 否'} |"
        )

    lines += [
        "",
        "## IC 为空 — 待后续制作",
        "",
        "输出文件：",
        f"- 清单 Excel：`{OUT_XLSX.relative_to(ROOT)}`",
        f"- 排期表副本（H 列标记 + IC 列高亮）：`{OUT_TIMELINE.relative_to(ROOT)}`",
        "",
    ]

    for lv in ("L3", "L4"):
        sub = [x for x in issues if x["level"] == lv]
        lines.append(f"### {lv}（共 {len(sub)} 本）")
        lines.append("")
        if not sub:
            lines.append("无。")
            lines.append("")
            continue
        lines.append("| Book# | S&S 标题 | Timeline 标题 | PPT | 正文 IP | 体裁 | 标题一致 |")
        lines.append("|-------|---------|--------------|-----|---------|------|---------|")
        for x in sorted(sub, key=lambda i: i["book_no"]):
            lines.append(
                f"| {x['book_no']} | {x['title_ss']} | {x['title_timeline']} | {x['ppt_status']} | "
                f"{x['ips_in_story']} | {x['genre'][:24] if x['genre'] else '—'} | "
                f"{'✓' if x['title_match'] else '✗'} |"
            )
        lines.append("")

    OUT_MD.write_text("\n".join(lines), encoding="utf-8")

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "IC待制作清单"
    ws_out.append([
        "级别", "Book#", "S&S标题", "Timeline标题", "IC现状", "PPT状态",
        "正文IP(参考)", "体裁", "主题", "标题一致", "备注",
    ])
    amber = PatternFill("solid", fgColor="FEF3C7")
    for x in sorted(issues, key=lambda i: (i["level"], i["book_no"])):
        note = "待分配IC / 未制作"
        if x["genre"] and "non" in x["genre"].lower():
            note += "；非虚构"
        if not x["title_match"]:
            note += "；标题与S&S不一致"
        ws_out.append([
            x["level"], x["book_no"], x["title_ss"], x["title_timeline"], x["ic"],
            x["ppt_status"], x["ips_in_story"], x["genre"], x["theme"],
            "Y" if x["title_match"] else "N", note,
        ])
        for col in (1, 2, 5, 11):
            ws_out.cell(ws_out.max_row, col).fill = amber
    wb_out.save(OUT_XLSX)

    mark_timeline_copy(issues)

    print(f"WROTE {OUT_MD}")
    print(f"WROTE {OUT_XLSX}")
    print(f"WROTE {OUT_TIMELINE}")
    print("\nSUMMARY:")
    for lv, total, miss, tm in summary:
        print(f"  {lv}: {total} books, {miss} IC empty (pending), {tm} title mismatches")
    print(f"\nTotal pending: {len(issues)}")


if __name__ == "__main__":
    main()
