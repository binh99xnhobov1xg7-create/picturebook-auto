"""从钉钉自动拉取 S&S 大纲并重建 references/syllabus/syllabus.json。

设计：dws 负责取数，_build_syllabus_snapshot 负责解析（复用既有列映射逻辑）。

用法：
  py scripts/sync_syllabus_from_dingtalk.py --dry-run     # 只下载+探测，不写 JSON
  py scripts/sync_syllabus_from_dingtalk.py               # 全量同步
  py scripts/sync_syllabus_from_dingtalk.py --notify      # 同步后钉钉群通知

环境变量（可选，默认用下方常量）：
  DINGTALK_L02_NODE_ID
  DINGTALK_L36_SPACE_ID / DINGTALK_L36_FILE_ID
  DINGTALK_REQUIREMENTS_NODE_ID
"""
from __future__ import annotations

import argparse
import hashlib
import json
import os
import shutil
import subprocess
import sys
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

REPO = Path(__file__).resolve().parents[1]
CACHE = REPO / "references" / "syllabus" / "_dingtalk_cache"
STATE = CACHE / "sync_state.json"
OUT_JSON = REPO / "references" / "syllabus" / "syllabus.json"
REQ_MD = REPO / "references" / "dingtalk" / "requirements.md"

L02_NODE = os.getenv("DINGTALK_L02_NODE_ID", "lyQod3RxJK3eA167sdB30NAqJkb4Mw9r")
L36_SPACE = os.getenv("DINGTALK_L36_SPACE_ID", "28534639840")
L36_FILE = os.getenv("DINGTALK_L36_FILE_ID", "220219603562")
REQ_NODE = os.getenv("DINGTALK_REQUIREMENTS_NODE_ID", "7NkDwLng8ZMaj15pHaqGnz5jJKMEvZBY")
REQ_XLSX = REPO / "references" / "dingtalk" / "requirements.xlsx"


def _dws_bin() -> str:
    found = shutil.which("dws")
    if found:
        return found
    for candidate in (
        Path.home() / ".local" / "bin" / "dws.exe",
        Path.home() / ".local" / "bin" / "dws",
    ):
        if candidate.exists():
            return str(candidate)
    return "dws"


def _run(cmd: list[str], *, timeout: int = 600) -> subprocess.CompletedProcess[str]:
    if cmd and cmd[0] == "dws":
        cmd = [_dws_bin(), *cmd[1:]]
    print("+", " ".join(cmd))
    return subprocess.run(
        cmd, check=True, timeout=timeout, text=True, encoding="utf-8", errors="replace",
    )


def _run_json(cmd: list[str], *, timeout: int = 600) -> dict:
    if cmd and cmd[0] == "dws":
        cmd = [_dws_bin(), *cmd[1:], "--format", "json"]
    else:
        cmd = [*cmd, "--format", "json"]
    print("+", " ".join(cmd))
    out = subprocess.check_output(cmd, timeout=timeout, text=True, encoding="utf-8")
    return json.loads(out)


def _http_download(url: str, dst: Path) -> None:
    dst.parent.mkdir(parents=True, exist_ok=True)
    req = urllib.request.Request(url, headers={"User-Agent": "picturebook-auto/1.0"})
    with urllib.request.urlopen(req, timeout=300) as resp:
        dst.write_bytes(resp.read())


def _pick_download_url(payload: dict) -> str | None:
    for key in ("resourceUrl", "downloadUrl", "urlPreSignature", "url"):
        val = payload.get(key)
        if isinstance(val, str) and val.startswith("http"):
            return val
    urls = payload.get("resourceUrls")
    if isinstance(urls, list) and urls:
        first = urls[0]
        if isinstance(first, str) and first.startswith("http"):
            return first
        if isinstance(first, dict):
            for key in ("url", "resourceUrl", "downloadUrl"):
                val = first.get(key)
                if isinstance(val, str) and val.startswith("http"):
                    return val
    for wrap in ("content", "result", "data"):
        nested = payload.get(wrap)
        if isinstance(nested, dict):
            found = _pick_download_url(nested)
            if found:
                return found
    return None


def _ensure_xlsx(path: Path) -> None:
    if not path.exists() or path.stat().st_size < 1024:
        raise RuntimeError(f"下载失败或文件过小: {path}")


def _file_hash(p: Path) -> str:
    return hashlib.sha256(p.read_bytes()).hexdigest()


def _dws_doc_info(node: str) -> dict:
    return _run_json(["dws", "doc", "info", "--node", node])


def _download_doc_file(node: str, dst: Path) -> None:
    """DOCUMENT/xlsx 节点：dws doc download，必要时跟预签名 URL。"""
    cmd = [_dws_bin(), "doc", "download", "--node", node, "--output", str(dst)]
    print("+", " ".join(cmd))
    proc = subprocess.run(
        cmd,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        timeout=600,
    )
    if dst.exists() and dst.stat().st_size >= 1024:
        return

    payload: dict | None = None
    for chunk in (proc.stdout, proc.stderr):
        chunk = (chunk or "").strip()
        if chunk.startswith("{"):
            try:
                payload = json.loads(chunk)
                break
            except json.JSONDecodeError:
                pass

    if payload:
        url = _pick_download_url(payload)
        if url:
            _http_download(url, dst)
            return

    if proc.returncode != 0:
        err = (proc.stderr or proc.stdout or "").strip()
        raise subprocess.CalledProcessError(proc.returncode, cmd, output=err)

    raise RuntimeError(f"doc download 未落盘: {dst}")


def fetch_l02_xlsx(dst: Path) -> Path:
    """L0-L2：alidocs 节点 → 本地 xlsx。"""
    info = _dws_doc_info(L02_NODE)
    ext = (info.get("extension") or "").lower()
    if ext == "axls":
        _run(["dws", "sheet", "export", "--node", L02_NODE, "--output", str(dst)])
    else:
        _download_doc_file(L02_NODE, dst)
    _ensure_xlsx(dst)
    return dst


def fetch_l36_xlsx(dst: Path) -> Path:
    """L3-L6：钉盘 xlsx。"""
    _run([
        "dws", "drive", "download",
        "--space-id", L36_SPACE,
        "--file-id", L36_FILE,
        "--output", str(dst),
    ])
    return dst


def fetch_requirements_snapshot(dst_md: Path, dst_xlsx: Path) -> Path | None:
    """第三份文档：Timeline 需求表（axls）→ xlsx + 索引 md。"""
    try:
        info = _dws_doc_info(REQ_NODE)
    except (subprocess.CalledProcessError, json.JSONDecodeError) as e:
        print("WARN: requirements doc info failed, skip:", e)
        return None

    ext = (info.get("extension") or "").lower()
    name = info.get("name") or "requirements"
    synced_at = datetime.now(timezone.utc).isoformat()

    try:
        if ext == "axls":
            _run(["dws", "sheet", "export", "--node", REQ_NODE, "--output", str(dst_xlsx)])
            _ensure_xlsx(dst_xlsx)
            sheets = _run_json(["dws", "sheet", "list", "--node", REQ_NODE])
            raw_sheets = sheets.get("sheets") or sheets.get("result")
            if not isinstance(raw_sheets, list) and isinstance(sheets, list):
                raw_sheets = sheets
            if not isinstance(raw_sheets, list):
                raw_sheets = []
            sheet_names = [
                s.get("name") or s.get("sheetName") or str(s)
                for s in raw_sheets
                if isinstance(s, dict)
            ]
            body = (
                f"# {name}\n\n"
                f"- 同步时间: {synced_at}\n"
                f"- 节点: `{REQ_NODE}`\n"
                f"- 类型: axls → `{dst_xlsx.name}`\n"
                f"- 工作表 ({len(sheet_names)}): {', '.join(sheet_names) or '(见 xlsx)'}\n\n"
                f"完整内容见同目录 `requirements.xlsx`。\n"
            )
            dst_md.write_text(body, encoding="utf-8")
            return dst_md

        if ext in ("adoc",) or str(info.get("contentType") or "").upper() == "ALIDOC":
            _run(["dws", "doc", "read", "--node", REQ_NODE, "--output", str(dst_md)])
            return dst_md if dst_md.exists() else None

        _download_doc_file(REQ_NODE, dst_xlsx)
        _ensure_xlsx(dst_xlsx)
        dst_md.write_text(
            f"# {name}\n\n- 同步时间: {synced_at}\n- 文件: `{dst_xlsx.name}`\n",
            encoding="utf-8",
        )
        return dst_md
    except (subprocess.CalledProcessError, RuntimeError, OSError) as e:
        print("WARN: requirements export failed, skip:", e)
        return None


def rebuild_syllabus(l02_xlsx: Path, l36_xlsx: Path) -> dict:
    """复用 _build_syllabus_snapshot 解析器，避免重复列映射。"""
    import openpyxl
    from _build_syllabus_snapshot import parse_l02, parse_l36, parse_sor_sheet

    books: dict[str, dict] = {}
    sor_strategies: dict[str, dict] = {}

    wb36 = openpyxl.load_workbook(l36_xlsx, data_only=True)
    for lvl in ("3", "4", "5", "6"):
        sheet = f"Level {lvl}"
        if sheet in wb36.sheetnames:
            books.update(parse_l36(wb36[sheet], lvl))

    wb02 = openpyxl.load_workbook(l02_xlsx, data_only=True)
    for lvl in ("0", "1", "2"):
        sheet = f"Level {lvl}"
        if sheet in wb02.sheetnames:
            books.update(parse_l02(wb02[sheet], lvl))
    if "SoR" in wb02.sheetnames:
        sor_strategies = parse_sor_sheet(wb02["SoR"])

    return {
        "_meta": {
            "source_l36": l36_xlsx.name,
            "source_l02": l02_xlsx.name,
            "book_count": len(books),
            "synced_at": datetime.now(timezone.utc).isoformat(),
            "sync_via": "dingtalk-workspace-cli",
        },
        "books": books,
        "sor_strategies": sor_strategies,
    }


def load_state() -> dict:
    if STATE.exists():
        return json.loads(STATE.read_text(encoding="utf-8"))
    return {}


def save_state(state: dict) -> None:
    STATE.parent.mkdir(parents=True, exist_ok=True)
    STATE.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")


def maybe_notify(changed: list[str]) -> None:
    if not changed:
        return
    try:
        sys.path.insert(0, str(REPO / "scripts"))
        from dingtalk_notify import send_dingtalk_markdown

        body = "### 大纲已自动同步\n\n" + "\n".join(f"- {c}" for c in changed)
        body += "\n\n@粟千雪 如有需求文档变更请确认 `references/dingtalk/requirements.md`"
        send_dingtalk_markdown("绘本 · 大纲同步", body)
    except Exception as e:
        print("notify skip:", e)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true", help="只下载到 cache，不写 syllabus.json")
    ap.add_argument("--notify", action="store_true", help="有变更时发钉钉群通知")
    ap.add_argument("--force", action="store_true", help="忽略 hash，强制重建")
    args = ap.parse_args()

    if _dws_bin() == "dws" and not shutil.which("dws"):
        print("ERROR: 请先安装并登录 dws (dws auth login)")
        return 1

    CACHE.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    l02_path = CACHE / f"l02_{ts}.xlsx"
    l36_path = CACHE / f"l36_{ts}.xlsx"
    req_md_path = CACHE / f"requirements_{ts}.md"
    req_xlsx_path = CACHE / f"requirements_{ts}.xlsx"

    fetch_l02_xlsx(l02_path)
    fetch_l36_xlsx(l36_path)
    fetch_requirements_snapshot(req_md_path, req_xlsx_path)

    state = load_state()
    changed: list[str] = []
    req_hash_path = req_xlsx_path if req_xlsx_path.exists() else req_md_path
    for label, p in [("L0-L2", l02_path), ("L3-L6", l36_path), ("requirements", req_hash_path)]:
        if not p.exists():
            continue
        h = _file_hash(p)
        if args.force or state.get(label) != h:
            changed.append(f"{label} ({p.name})")
            state[label] = h

    if not args.dry_run:
        REQ_MD.parent.mkdir(parents=True, exist_ok=True)
        if req_md_path.exists():
            REQ_MD.write_text(req_md_path.read_text(encoding="utf-8"), encoding="utf-8")
        if req_xlsx_path.exists():
            REQ_XLSX.write_bytes(req_xlsx_path.read_bytes())

    if not changed and not args.force:
        print("No changes detected, skip rebuild.")
        save_state(state)
        return 0

    if args.dry_run:
        print("DRY-RUN: downloaded to", CACHE, "| would rebuild:", changed)
        save_state(state)
        return 0

    payload = rebuild_syllabus(l02_path, l36_path)
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=1), encoding="utf-8")
    print("WROTE", OUT_JSON, "books=", payload["_meta"]["book_count"])

    sys.path.insert(0, str(REPO / "scripts"))
    try:
        from syllabus import load_syllabus
        load_syllabus.cache_clear()
    except Exception:
        pass

    save_state(state)
    if args.notify:
        maybe_notify(changed)
    return 0


if __name__ == "__main__":
    sys.exit(main())
