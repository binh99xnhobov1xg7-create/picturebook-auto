"""按最新规则批量跑 L4 IC 待制作 24 本（Book# 3,6,9…72）。

数据源：~/下载/VIPKID/大纲/Level 3-6  S&S.xlsx · Level 4 表
规则对齐 transcript 0a534f10：GPT 出图、10 岁 IP(mia_10/tommy_10)、20-25% 顶部场景留白、
L4 worksheet/RR 版式、官方 S&S 注入 Strategy/Skill/GO。
"""
from __future__ import annotations

import os
os.environ["PYTHONUNBUFFERED"] = "1"

import re
import sys
import time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

import openpyxl
from batch_runner import BatchItem, run_batch, set_image_concurrency
from config import OUTPUTS_DIR, resolve_image_backend

XLSX = Path.home() / "下载" / "VIPKID" / "大纲" / "Level 3-6  S&S.xlsx"
WANT = [3, 6, 9, 12, 15, 18, 21, 24, 27, 30, 33, 36, 39, 42, 45, 48, 51, 54, 57, 60, 63, 66, 69, 72]
OUT_ROOT = OUTPUTS_DIR / "L4_batch"


def _split_pages(col10: str) -> str:
    s = (col10 or "").strip()
    if not s:
        return ""
    parts = re.split(r"Page\s*\d+\s*[:：]?", s)
    pages = [re.sub(r"\s+", " ", p).strip() for p in parts if p.strip()]
    return "\n".join(pages)


def _cefr(raw) -> str:
    c = str(raw or "").strip()
    if c.upper().startswith("CEFR"):
        c = c[4:].strip()
    return c


def _load_items() -> list[BatchItem]:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Level 4"]
    rows: dict[int, int] = {}
    for r in range(2, ws.max_row + 1):
        bn = ws.cell(r, 1).value
        try:
            no = int(float(bn))
        except (TypeError, ValueError):
            continue
        if no in WANT:
            rows[no] = r
    wb.close()

    items: list[BatchItem] = []
    for bn in WANT:
        r = rows.get(bn)
        if r is None:
            print(f"[SKIP] book {bn} not in S&S", flush=True)
            continue
        title = str(ws.cell(r, 3).value or "").strip()
        genre = str(ws.cell(r, 2).value or "")
        fiction_type = (
            "non-fiction" if "非虚构" in genre or "nonfiction" in genre.lower() else "fiction"
        )
        cefr = _cefr(ws.cell(r, 24).value)
        story = _split_pages(str(ws.cell(r, 20).value or ""))
        if not story:
            story = str(ws.cell(r, 21).value or "").strip()
        theme = str(ws.cell(r, 6).value or "").strip()
        items.append(
            BatchItem(
                title=title,
                level="4",
                book_number=f"{bn:02d}",
                story=story,
                cefr=cefr,
                theme=theme,
                fiction_type=fiction_type,
            )
        )
    wb.close()
    return items


def main() -> None:
    items = _load_items()
    if not items:
        raise SystemExit("没有可跑的 L4 书目。")

    OUT_ROOT.mkdir(parents=True, exist_ok=True)
    set_image_concurrency(2)
    print(
        f"=== L4 IC BATCH: {len(items)} books | backend={resolve_image_backend('4')} "
        f"| out={OUT_ROOT} ===",
        flush=True,
    )
    for it in items:
        print(f"  queued #{it.book_number} {it.title} ({it.fiction_type})", flush=True)

    def cb(done: int, total: int, r) -> None:
        msg = (
            f"[{done}/{total}] #{r.item.book_number} {r.item.title} | "
            f"status={r.status} eval={r.eval_level or '-'} review={r.needs_human_review} "
            f"{r.elapsed_s:.0f}s"
        )
        if r.error:
            msg += " | ERR " + r.error[:120].replace("\n", " ")
        print(msg, flush=True)

    t0 = time.time()
    summary = run_batch(
        items,
        out_root=OUT_ROOT,
        concurrency=1,
        image_concurrency=2,
        mock=False,
        resume=True,
        progress_cb=cb,
    )
    print(
        f"\n=== DONE L4_IC_BATCH ok={summary['ok']} failed={summary['failed']} "
        f"need_review={summary.get('need_review', 0)} elapsed={time.time()-t0:.0f}s ===",
        flush=True,
    )
    print(f"OUT: {OUT_ROOT}", flush=True)
    print(f"LOG: {summary.get('log_path')}", flush=True)
    print("DONE_L4_IC_BATCH", flush=True)


if __name__ == "__main__":
    main()
