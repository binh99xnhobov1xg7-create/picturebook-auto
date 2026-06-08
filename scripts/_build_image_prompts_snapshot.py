# -*- coding: utf-8 -*-
"""把官方《全级别每课Prompt.xlsx》解析为 references/syllabus/image_prompts.json。

每本书一条：level / book_label / title / full_text / pages[]（尽量切出的逐页画面要求）。
出图时按 (level, title) 检索，作为"权威参考"注入逐页生成（不替换 IP 锁脸与画风）。
"""
from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl

SRC = Path(r"C:\Users\Jered\下载\VIPKID\02. 绘本PPT制作说明\全级别每课Prompt.xlsx")
OUT = Path(__file__).resolve().parent.parent / "references" / "syllabus" / "image_prompts.json"


def _clean(v) -> str:
    if v is None:
        return ""
    return str(v).replace("\r", "").replace("\u200b", "").replace("\ufeff", "").strip()


def _norm_title(s: str) -> str:
    s = (s or "").lower()
    s = re.sub(r"[^a-z0-9]+", "", s)
    return s


_TITLE_BRACKET = re.compile(r"《\s*(.+?)\s*》")
_TITLE_HEADING = re.compile(r"^#{1,6}\s*#?\s*(.+?)\s*$", re.M)
_BOOK_TITLE = re.compile(r"Book\s*\d+\s*[：:]?\s*[《\"]?\s*([A-Za-z][^《》\"\n]{1,45}?)\s*[》\"]?\s*$", re.I | re.M)
# 英文书名出现在「完整生图 Prompt」之前（如 "春日绘本：Spring Days 完整生图 Prompt"）
_TITLE_BEFORE = re.compile(r"([A-Za-z][A-Za-z0-9'’,!?\.\- ]{2,45}?)\s*(?:完整生图|完整\s*生图|生图\s*Prompt|完整\s*Prompt|的\s*AI\s*生图)", re.I)
_LEVEL_PREFIX = re.compile(r"^\s*(?:#+\s*)?(?:Level\s*\d+\s*)?(?:Book\s*\d+\s*)?[：:]?\s*", re.I)


def _clean_title(t: str) -> str:
    t = t.strip().strip("*#").strip()
    t = _LEVEL_PREFIX.sub("", t).strip()
    t = re.split(r"\s*(完整生图|完整|生图|Prompt|的\s*AI)", t)[0].strip()
    return t.strip(" :：-—")


def _extract_title(text: str) -> str:
    # 1) 《Title》
    m = _TITLE_BRACKET.search(text)
    if m:
        t = _clean_title(m.group(1))
        if t:
            return t
    # 2) Book N《Title》/ Book N: Title
    m = _BOOK_TITLE.search(text)
    if m:
        t = _clean_title(m.group(1))
        if t:
            return t
    # 3) "XXX：English Title 完整生图 Prompt"
    m = _TITLE_BEFORE.search(text)
    if m:
        t = _clean_title(m.group(1))
        if t and re.search(r"[A-Za-z]", t):
            return t
    # 4) markdown 标题里第一行像书名的（英文为主），去掉 Level/Book 前缀
    for mm in _TITLE_HEADING.finditer(text):
        cand = _clean_title(mm.group(1))
        if cand and re.search(r"[A-Za-z]", cand) and len(cand) < 60 \
           and not cand.startswith(("关键", "故事", "通用", "全局", "画面", "封面")):
            return cand
    return ""


# 逐页切分：常见标记
_PAGE_MARKERS = re.compile(
    r"(封面|内封|第\s*[0-9一二三四五六七八九十]+\s*页|P\s*[0-9]+\b|Page\s*[0-9]+|"
    r"跨页|Spread|第[0-9]+跨页)", re.I,
)
# 画面要求块
_SCENE_REQ = re.compile(r"(?:画面要求|画面|场景)[：:]\s*(.+)")


def _split_pages(text: str) -> list[dict]:
    """尽量把整段 prompt 切成 [{marker, text, scene}]。失败则返回 []。"""
    matches = list(_PAGE_MARKERS.finditer(text))
    if len(matches) < 2:
        return []
    pages = []
    for i, m in enumerate(matches):
        start = m.start()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        seg = text[start:end].strip()
        sm = _SCENE_REQ.search(seg)
        scene = ""
        if sm:
            scene = sm.group(1).strip()
            scene = re.split(r"\n\s*>?\s*\**Prompt", scene)[0].strip()[:300]
        pages.append({
            "marker": m.group(1).strip(),
            "scene": scene,
            "text": seg[:1200],
        })
    return pages


def parse_sheet(ws, level: str) -> list[dict]:
    out = []
    for row in ws.iter_rows(values_only=True):
        cells = [_clean(c) for c in row]
        cells = [c for c in cells if c]
        if not cells:
            continue
        label = cells[0]
        # 取最长的一格作为 prompt 正文
        body_cells = cells[1:] if len(cells) > 1 else cells
        prompt = max(body_cells, key=len) if body_cells else ""
        if len(prompt) < 60:  # 太短的多半是表头/说明
            continue
        title = _extract_title(prompt)
        if not title:
            continue
        out.append({
            "level": level,
            "book_label": label if label != prompt else "",
            "title": title,
            "norm_title": _norm_title(title),
            "pages": _split_pages(prompt),
            "full_text": prompt[:8000],
        })
    return out


def main():
    wb = openpyxl.load_workbook(str(SRC), data_only=True)
    entries = []
    for sn in wb.sheetnames:
        m = re.fullmatch(r"[Ll]?\s*([0-6])", sn.strip())
        if not m:
            continue
        level = m.group(1)
        rows = parse_sheet(wb[sn], level)
        entries.extend(rows)
        print(f"sheet {sn}: {len(rows)} books")
    wb.close()
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(entries, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"WROTE {OUT}  total={len(entries)}")
    # 抽样
    for e in entries[:6]:
        print(f"  L{e['level']} [{e['book_label']}] {e['title']}  pages={len(e['pages'])}")


if __name__ == "__main__":
    main()
