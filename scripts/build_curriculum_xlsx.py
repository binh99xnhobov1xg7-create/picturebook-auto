# -*- coding: utf-8 -*-
"""生成「0-6 课程对标总表」Excel（销售/教研/家长可直接打开分享）。

设计口径（用户 2026-06-07 拍板）：
- 级别配色 = 练习册同款（config.BRAND_COLORS，按级别一色）。
- 删掉「主角形象年龄」「听说读写四项百分比」。
- 加入：学生实际年龄 / 国内年级 / 美国年级 / 剑桥少儿(YLE) / 剑桥主体(KET-PET) / AR 值 / 词汇量。
- 考察方向只标「以听说为中心 / 以读写为中心」，不拆百分比。
- 核心目标用概括性能力动作词（参考 RAZ / 牛津分级外化方式）。

2026-06-08 对标 VIPKID 主修课 / 北美外教阅读课物料后补充（用户确认）：
- 新增「阶段与培养目标」段：阅读阶段(夯实基础→进阶→流利→自主) / 语言发展 / 素质培养。
- 对标基准新增：TOEFL Primary·Junior(仅 L3+ 参考) / RAZ 阅读体系(与 AR 并存) / 累计阅读字数。
- 语言知识体系新增：阅读文体(体裁)；学习重点新增：阅读策略(predict/visualize/summarize/retell)。
- 不引入 IELTS / TOEFL-iBT / FCE(B2) / 课程节数：超出 0-6(4-11 岁, 天花板 B1)体系或不适用绘本生成产品。

权威值（CEFR / 蓝思 / 解码句法 / 核心词/本 / 词汇主题 / 阅读技能 / 产出）来自官方 S&S + TG SOP，
与 level_profiles 同源；对标参考值（学生年龄 / 年级 / 剑桥考试 / AR / 累计词汇量）依 CEFR 推导。
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
try:
    from config import OUTPUTS_DIR, brand_color_hex
except Exception:  # pragma: no cover
    OUTPUTS_DIR = Path(__file__).resolve().parent.parent / "outputs"
    _FALLBACK = {"0": "#5E9F49", "1": "#F18200", "2": "#54C2F0", "3": "#E94653",
                 "4": "#00B0C4", "5": "#E95283", "6": "#0677B7"}

    def brand_color_hex(level: str) -> str:  # type: ignore
        return _FALLBACK.get(str(level), "#E95283")


OUT = OUTPUTS_DIR / "_framework" / "课程对标总表_L0-L6.xlsx"

LEVEL_KEYS = ["0", "1", "2", "3", "4", "5", "6"]
LEVELS = [f"L{k}" for k in LEVEL_KEYS]

# 表头副信息（每级别表头下方两行）：学生年龄 / 国内年级
HEADER_AGE = ["4–5 岁", "5–6 岁", "6–7 岁", "7–8 岁", "8–9 岁", "9–10 岁", "10–11 岁"]
HEADER_GRADE = ["幼儿园", "学前", "一年级", "二年级", "三年级", "四年级", "五年级"]

# 级别配色（练习册同款）
LEVEL_HEX = [brand_color_hex(k) for k in LEVEL_KEYS]

# 标了 ᴿ 的维度 = 对标参考（依 CEFR 推导，可校准）；其余为权威值
SUGGEST = {
    "学生实际年龄", "国内年级", "美国对应年级", "剑桥少儿 (YLE)", "剑桥主体 (KET/PET)",
    "TOEFL 测评", "RAZ 阅读体系", "AR 值 (ATOS)", "词汇量 (累计)", "累计阅读字数",
    "语言发展", "素质培养", "核心目标", "考察方向",
}

DATA: dict[str, dict[str, list[str]]] = {
    "① 阶段与培养目标": {
        "阅读阶段": ["夯实基础", "夯实基础", "夯实基础", "进阶提升", "进阶提升",
                 "流利阅读", "自主阅读"],
        "语言发展": ["兴趣启蒙", "兴趣启蒙", "基础夯实", "基础夯实", "技能提升",
                 "技能提升", "综合飞跃"],
        "素质培养": ["构建国际文化意识", "构建国际文化意识", "提升交流沟通能力",
                 "提升交流沟通能力", "激发内在创新动力", "激发内在创新动力",
                 "培养独立批判思维"],
    },
    "② 对标基准": {
        "学生实际年龄": HEADER_AGE,
        "国内年级": ["幼儿园", "学前 / 幼小衔接", "小学一年级", "小学二年级",
                 "小学三年级", "小学四年级", "小学五年级"],
        "美国对应年级": ["Pre-K", "K", "G1", "G2", "G3", "G4", "G5"],
        "欧标 CEFR": ["Pre-A1", "Pre-A1→A1", "A1", "A1+", "A2", "A2+", "B1"],
        "剑桥少儿 (YLE)": ["预备级", "Starters", "Starters–Movers", "Movers",
                       "Flyers", "Flyers", "—"],
        "剑桥主体 (KET/PET)": ["—", "—", "—", "—", "—", "KET (A2)", "KET–PET (B1)"],
        "TOEFL 测评": ["—", "—", "—", "TOEFL Primary Step 1",
                     "TOEFL Primary Step 1–2", "TOEFL Primary Step 2",
                     "TOEFL Primary Step 2 / Junior 入门"],
        "RAZ 阅读体系": ["aa–A", "A–B", "C–D", "E–G", "H–J", "K–M", "N–P"],
        "AR 值 (ATOS)": ["0.5–1.0", "0.8–1.3", "1.0–1.8", "1.8–2.5",
                      "2.5–3.5", "3.5–4.5", "4.5–5.5"],
        "蓝思 Lexile": ["10–200L", "10–200L", "10–200L", "210–400L",
                      "410–600L", "410–600L", "600L+"],
        "词汇量 (累计)": ["~150", "~300", "~500", "~800", "~1000", "~1500", "~2500"],
        "累计阅读字数": ["~1,000", "~2,500", "~5,000", "~9,000", "~16,000",
                    "~28,000", "~45,000"],
        "单本正文字数": ["9–33 词", "15–40 词", "20–50 词", "70–85 词",
                    "90–100 词", "130–167 词", "200–210 词"],
    },
    "③ 语言知识体系": {
        "解码 / 构词": ["音素意识", "自然拼读·解码", "自然拼读·规律复习",
                    "自然拼读·意识", "自然拼读·意识(更难)", "构词法", "构词法(进阶)"],
        "句法重点": ["单词标签 + 句型框架", "可解码句型 + 顺序词", "故事语法(问题/解决) + 词组句",
                  "目标句型(时态/介词/从句雏形)", "关系从句/时间从句/定义句",
                  "复合句/连接词/说明议论句式", "长难句/被动语态/复杂句式"],
        "核心词汇量 (每本)": ["8 词", "6 词", "5 词", "4 词", "4 词", "5 词", "5 词"],
        "词汇主题": ["颜色/数字/形状/食物/动物/家庭", "宠物/日常/动作/天气/社区",
                  "故事/解决问题/友谊/探索/自然", "时间/身体/动物/五感/季节",
                  "世界地形/海洋/雨林/栖息地", "友谊社交/品格/社会科学议题",
                  "生存与适应/科学社会议题"],
        "阅读文体 (体裁)": ["概念/标签书 · 童谣", "可解码故事 · 生活记叙",
                       "记叙故事(问题-解决)", "记叙文 + 简单说明文",
                       "说明文(科学) + 记叙文", "说明文/议论雏形 + 人物",
                       "说明/议论文 + 人物传记"],
    },
    "④ 学习重点与考察": {
        "核心目标": ["听懂 · 会指认", "拼读 · 敢开口", "读懂小故事", "自主读 · 会提取",
                 "抓细节 · 会归纳", "会推理 · 会比较", "能思辨 · 会评价"],
        "考察方向": ["以听说为中心", "以听说为中心", "以听说为中心(读起步)",
                 "听说 → 读写过渡", "以读写为中心", "以读写为中心", "以读写为中心"],
        "阅读策略": ["看图预测 · 联系生活", "预测 · 提问与回答",
                  "预测 · 可视化 · 提问回答", "预测 · 联系已知 · 可视化 · 提问回答",
                  "预测 · 可视化 · 提问回答 · 归纳总结", "归纳总结 · 复述 · 修正预测",
                  "综合运用全部策略 · 自我监控复述"],
        "阅读技能": ["看图命名识别", "顺序词复述", "故事语法(问题/解决)",
                  "分层理解·按体裁读·GO 整理", "主旨细节·事实观点·读懂从句",
                  "推理·比较对比·作者意图", "评价·论证·跨文本综合"],
        "思维层级": ["回忆", "回忆→简单推断", "推断显现", "推断为主·起步分析",
                  "推断 + 分析", "分析为主", "分析 / 评价"],
    },
    "⑤ 能力 · 题型": {
        "Worksheet 主力题型": ["看图连线/配对、圈词、描红", "看图选词、拼读补全、排序、连线",
                          "词义配对、单词补全、看图选句、一句话", "词义匹配、阅读单选、GO 图形组织器",
                          "词义匹配、阅读单选、时态改写、主旨细节 GO",
                          "阅读分析单选、比较对比 GO、句子合并、段落写作",
                          "批判性阅读题、论证证据、跨文本综合、结构化写作"],
        "阅读能力": ["看图认词/跟读/指读", "解码 CVC/跟读/顺序词复述", "词组流畅朗读/追踪问题-解决",
                  "分层理解/按体裁读/用 GO 整理", "主旨与细节/事实观点/读懂从句",
                  "推理/比较对比/作者意图/综合归纳", "评价/论证/跨文本综合"],
        "写作能力": ["描红/连线(前书写)", "抄写单词/补全字母", "写一句话回应",
                  "目标句型造句/GO 整理", "用从句写句/整理主旨细节",
                  "用复合句写较长回应/短段落", "结构化长段落/论证表达"],
    },
    "⑥ 学完产出": {
        "得到（成果物）": ["能用句型框架说出一页", "能独立解码并跟读一本绘本",
                     "能读懂问题-解决结构的简单故事", "能读懂~80词短文并完成 GO",
                     "能读懂~95词含从句短文并完成 GO", "能读懂~150词较长文本并完成高阶 GO/写作",
                     "能读懂~200词说明/议论文本并完成批判性任务"],
        "学到（知识/技能）": ["主题词命名、字母-声音意识、印刷概念", "短元音/首尾辅音解码、顺序词复述",
                       "拼读自动化、问题/解决结构、新词", "语境内学词、指定阅读策略、目标句型",
                       "关系/时间从句、主旨细节、事实观点", "构词法、复合句、推理与比较对比",
                       "学术构词、长难句、批判性阅读与论证"],
        "练到（迁移应用）": ["看图口头造句、跟读指读", "解码新词、复述顺序、找高频词",
                       "词组流畅朗读、追踪结构、写一句话", "分层答题、用 GO 整理、按句型表达",
                       "推断分析题、用从句写句、整理主旨", "分析评价题、综合归纳、用复合句写回应",
                       "评价论证题、跨文本综合、写结构化长段落"],
    },
}


# ---------------- 样式工具 ----------------
def _lighten(hex_: str, factor: float) -> str:
    """把颜色按 factor(0~1) 向白色混合，得到浅色底。"""
    h = hex_.lstrip("#")
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    r = int(r + (255 - r) * factor)
    g = int(g + (255 - g) * factor)
    b = int(b + (255 - b) * factor)
    return f"{r:02X}{g:02X}{b:02X}"


def level_header_hex(i: int) -> str:
    return LEVEL_HEX[i].lstrip("#")


def level_tint_hex(i: int) -> str:
    return _lighten(LEVEL_HEX[i], 0.82)


INK = "1F2937"
WHITE = "FFFFFF"
HEADER_BG = "F47332"
SECTION_BG = "374151"
DIM_BG = "F3F4F6"

thin = Side(style="thin", color="D1D5DB")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def _fill(hex_):
    return PatternFill("solid", fgColor=hex_)


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "0-6 课程对标总表"
    n_cols = 1 + len(LEVELS)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    t = ws.cell(1, 1, "VIPKID Dino · Levels 0–6 课程对标总表")
    t.font = Font(name="微软雅黑", size=16, bold=True, color=WHITE)
    t.fill = _fill(HEADER_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    s = ws.cell(2, 1, "一张表看懂每个级别：年龄学龄 · 欧标/剑桥/AR/蓝思对标 · 学习重点与考察 · 题型能力 · 学完产出")
    s.font = Font(name="微软雅黑", size=10, color="6B7280")
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # 表头（级别 + 年龄 + 年级，三行合一格，按级别配色）
    hr = 3
    c0 = ws.cell(hr, 1, "维度 \\ 级别")
    c0.font = Font(name="微软雅黑", size=11, bold=True, color=WHITE)
    c0.fill = _fill(SECTION_BG)
    c0.alignment = Alignment(horizontal="center", vertical="center")
    c0.border = BORDER
    for i, lv in enumerate(LEVELS):
        cell = ws.cell(hr, 2 + i, f"{lv}\n{HEADER_AGE[i]} · {HEADER_GRADE[i]}")
        cell.font = Font(name="微软雅黑", size=12, bold=True, color=WHITE)
        cell.fill = _fill(level_header_hex(i))
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[hr].height = 40

    r = hr + 1
    for section, dims in DATA.items():
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
        sc = ws.cell(r, 1, section)
        sc.font = Font(name="微软雅黑", size=11, bold=True, color=WHITE)
        sc.fill = _fill(SECTION_BG)
        sc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        sc.border = BORDER
        ws.row_dimensions[r].height = 22
        r += 1

        for dim, vals in dims.items():
            label = dim + ("  ᴿ" if dim in SUGGEST else "")
            dcell = ws.cell(r, 1, label)
            dcell.font = Font(name="微软雅黑", size=10, bold=True, color=INK)
            dcell.fill = _fill(DIM_BG)
            dcell.alignment = Alignment(horizontal="left", vertical="center",
                                        wrap_text=True, indent=1)
            dcell.border = BORDER
            for i, v in enumerate(vals):
                cell = ws.cell(r, 2 + i, v)
                cell.font = Font(name="微软雅黑", size=10, color=INK)
                cell.fill = _fill(level_tint_hex(i))
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                cell.border = BORDER
            ws.row_dimensions[r].height = 36
            r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
    f1 = ws.cell(r, 1, "权威值：欧标 CEFR / 蓝思 / 解码句法 / 核心词(本) / 词汇主题 / 阅读技能 / 学完产出 —— 来自官方 S&S 大纲 + TG_SOP，与系统 level_profiles 同源。")
    f1.font = Font(name="微软雅黑", size=9, color="6B7280")
    f1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
    f2 = ws.cell(r, 1, "标 ᴿ = 对标参考（学生年龄 / 国内·美国年级 / 剑桥 YLE·KET·PET / TOEFL / RAZ / AR 值 / 累计词汇量 / 累计阅读字数 / 语言发展 / 素质培养 / 核心目标 / 考察方向）：依 CEFR 推导，供校准，可按学情调整。")
    f2.font = Font(name="微软雅黑", size=9, color="6B7280")
    f2.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    ws.column_dimensions["A"].width = 20
    for i in range(len(LEVELS)):
        ws.column_dimensions[get_column_letter(2 + i)].width = 22

    ws.freeze_panes = "B4"

    OUT.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(OUT)
        print("WROTE", OUT)
    except PermissionError:
        # 目标被 Excel 占用：存一个带时间戳的备用名，避免直接失败
        import time
        alt = OUT.with_name(f"{OUT.stem}_{time.strftime('%H%M%S')}.xlsx")
        wb.save(alt)
        print("LOCKED, WROTE FALLBACK", alt)


if __name__ == "__main__":
    build()
