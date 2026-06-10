"""导出 Book54《Try Something New》提示词 + 抽取数据 + 当前代码 Mia 形象锁复核(无 API)。

输出: scripts/_book54_prompts_export.txt
- A 段: 源大纲抽取数据(直接读 Excel, 不调 LLM)。
- B 段: 实际出图所用整本提示词(复制成书时落盘的 image_prompts.txt)。
- C 段: 当前代码下 Mia 形象锁定 / HAIR LOCK 的真实输出(证明文字已是 half-up)。
"""
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
import openpyxl

BOOK = 54
XLSX = Path.home() / "下载" / "VIPKID" / "大纲" / "Level 3-6  S&S.xlsx"
ONDISK = Path(r"D:\picturebook_outputs\L3_batch\Level 3_Book54_Try_Something_New\image_prompts.txt")
OUT = Path(__file__).resolve().parent / "_book54_prompts_export.txt"


def _split_pages(col10: str) -> list[str]:
    s = (col10 or "").strip()
    if not s:
        return []
    parts = re.split(r"Page\s*\d+\s*[:：]?", s)
    return [re.sub(r"\s+", " ", p).strip() for p in parts if p.strip()]


def _cefr(raw) -> str:
    c = str(raw or "").strip()
    return c[4:].strip() if c.upper().startswith("CEFR") else c


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Level 3"]
    row = None
    for r in range(2, ws.max_row + 1):
        bn = ws.cell(r, 1).value
        if isinstance(bn, (int, float)) and int(bn) == BOOK:
            row = r
            break
    if row is None:
        raise SystemExit(f"Book {BOOK} not found in Excel")

    title = str(ws.cell(row, 4).value or "").strip()
    genre = str(ws.cell(row, 2).value or "")
    fiction_type = "non-fiction" if ("非虚构" in genre or "nonfiction" in genre.lower()) else "fiction"
    cefr = _cefr(ws.cell(row, 14).value)
    theme = str(ws.cell(row, 3).value or "").strip()
    pages = _split_pages(ws.cell(row, 10).value)
    if not pages:
        pages = [str(ws.cell(row, 11).value or "").strip()]

    # 当前代码: Mia 各档形象锁定 / HAIR LOCK 真实输出(无 API)
    from cn_prompt_builder import _make_protagonist_entry, MIA_HAIR_LOCK, MIA_HAIR_NEG
    from character_registry import get_description as reg_desc
    from ip_library import resolve_name_to_ip

    lines = []
    w = lines.append
    w(f"# Book{BOOK}《{title}》提示词 + 抽取数据导出")
    w(f"# 源 Excel: {XLSX}")
    w(f"# 生成方式: 直接读 Excel 抽取数据 + 复制成书时 image_prompts.txt + 当前代码无 API 复核")
    w("")
    w("=" * 70)
    w("A 段 · 源大纲抽取数据 (Excel)")
    w("=" * 70)
    w(f"book_number : {BOOK}")
    w(f"title       : {title}")
    w(f"genre(raw)  : {genre}")
    w(f"fiction_type: {fiction_type}")
    w(f"cefr        : {cefr}")
    w(f"theme       : {theme}")
    w(f"页数(正文)  : {len(pages)}")
    w("--- 逐页正文(抽取) ---")
    for i, p in enumerate(pages, start=1):
        w(f"[Page {i}] {p}")
    w("")

    w("=" * 70)
    w("C 段 · 当前代码 Mia 形象锁定 / HAIR LOCK 真实输出 (无 API, 证明文字逻辑)")
    w("=" * 70)
    w(f"[cn_prompt_builder.MIA_HAIR_LOCK]\n{MIA_HAIR_LOCK}")
    w(f"\n[cn_prompt_builder.MIA_HAIR_NEG]\n{MIA_HAIR_NEG}")
    w("")
    for age in (8, 10, 12):
        ent = _make_protagonist_entry("mia", age)
        ip = resolve_name_to_ip("Mia", age)
        w(f"--- Mia age {age} ---")
        w(f"  ip_library 取图: {ip.image_path if ip else None}")
        w(f"  【Mia 形象锁定】description_cn(当前代码, 出现在每页提示词):")
        w(f"    {ent.get('description_cn') if ent else None}")
        w(f"  cast_lock(character_registry 英文, 出现在修图锁):")
        w(f"    {reg_desc('mia', age)}")
        w("")

    w("=" * 70)
    w("B 段 · 实际出图所用整本提示词 (成书时落盘 image_prompts.txt 原文)")
    w("=" * 70)
    if ONDISK.exists():
        w(f"[来源] {ONDISK}")
        w("[注意] 该文件为成书时落盘, 若早于 2026-06-09 HAIR LOCK 修复, Mia 发型文字可能仍为旧版马尾描述(见报告说明)。")
        w("-" * 70)
        w(ONDISK.read_text(encoding="utf-8"))
    else:
        w(f"[缺失] 未找到 {ONDISK}")

    OUT.write_text("\n".join(lines), encoding="utf-8")
    print("WROTE", OUT)
    print("pages:", len(pages), "| title:", title)


if __name__ == "__main__":
    main()
