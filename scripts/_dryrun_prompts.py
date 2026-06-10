# -*- coding: utf-8 -*-
"""DRY-RUN：只构建并打印某本书各页最终出图 prompt（prompts_only=True，绝不调图片 API）。

用途：验证"角色外观锁不再夹带场景/情节句"的修复（Book63 根因）。
用法：
  py scripts/_dryrun_prompts.py 63
  py scripts/_dryrun_prompts.py 63 57 66     # 可一次多本
文本抽取仍走真实 LLM（零出图成本）；锚图不生成。
"""
import os
import re
import sys
import tempfile
from pathlib import Path

sys.path.insert(0, "scripts")
import openpyxl
from batch_runner import BatchItem, run_one

XLSX = Path.home() / "下载" / "VIPKID" / "大纲" / "Level 3-6  S&S.xlsx"


def _split_pages(col10: str) -> str:
    s = (col10 or "").strip()
    if not s:
        return ""
    parts = re.split(r"Page\s*\d+\s*[:：]?", s)
    pages = [re.sub(r"\s+", " ", p).strip() for p in parts if p.strip()]
    return "\n".join(pages)


def _cefr(raw) -> str:
    c = str(raw or "").strip()
    if c.upper().startswith("CEFR"):
        c = c[4:].strip()
    return c


def _load_item(bn: int) -> BatchItem:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Level 3"]
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, (int, float)) and int(v) == bn:
            title = str(ws.cell(r, 4).value or "").strip()
            genre = str(ws.cell(r, 2).value or "")
            ft = "non-fiction" if "非虚构" in genre or "nonfiction" in genre.lower() else "fiction"
            cefr = _cefr(ws.cell(r, 14).value)
            story = _split_pages(ws.cell(r, 10).value) or str(ws.cell(r, 11).value or "").strip()
            theme = str(ws.cell(r, 3).value or "").strip()
            return BatchItem(title=title, level="3", book_number=str(bn),
                             story=story, cefr=cefr, theme=theme, fiction_type=ft)
    raise SystemExit(f"[SKIP] book {bn} not found in 'Level 3' sheet")


def main():
    if len(sys.argv) < 2:
        raise SystemExit("用法: py scripts/_dryrun_prompts.py <book_number> [more...]")
    bns = [int(x) for x in sys.argv[1:]]
    env_root = os.getenv("DRYRUN_OUT")
    tmp_root = Path(env_root) if env_root else Path(tempfile.mkdtemp(prefix="pb_dryrun_"))
    tmp_root.mkdir(parents=True, exist_ok=True)
    print(f"[DRYRUN_OUT] {tmp_root}", flush=True)
    for bn in bns:
        item = _load_item(bn)
        print(f"\n########## DRY-RUN {bn} {item.title} ##########", flush=True)
        run_one(item, tmp_root, resume=False, prompts_only=True)
        pfile = tmp_root / item.name_prefix / "image_prompts.txt"
        print(f"[PROMPT_FILE] {bn} -> {pfile}", flush=True)
        if pfile.exists():
            print(pfile.read_text(encoding="utf-8"), flush=True)
        print(f"########## END {bn} ##########", flush=True)


if __name__ == "__main__":
    main()
